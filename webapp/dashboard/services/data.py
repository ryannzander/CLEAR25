"""
Station data loading: reads Excel regression files and coordinate sheets.

Data sources (per .cursorrules):
- Research: data/LAPTOP TSF 2026/07. The 4 Cities - Regression formulas and alert network stations/{City}/
- Legacy fallback: data/{City}_PM25_EWS_Regression.xlsx (sheets: Included Stations, All Stations Data)
- Bundled JSON: dashboard/services/bundled_stations.json when Excel data is missing or yields no rows
"""

import json
import os
from pathlib import Path

from django.conf import settings

DATA_DIR = settings.DATA_DIR
CONFIG_PATH = os.path.join(DATA_DIR, "config.json")
RESEARCH_DATA_BASE = getattr(settings, "RESEARCH_DATA_BASE", None)
NAPS_STATIONS_PATH = getattr(settings, "NAPS_STATIONS_PATH", None)

# Station IDs to exclude (too far from target city to be useful)
EXCLUDED_STATION_IDS = {"50308", "50310", "50314", "50313", "55702"}

# Toronto reference stations (methodology Section 2.2): use highest of 60430, 60410 for validation
TORONTO_NAPS_REFERENCE_IDS = {"60430", "60410"}

# Thunder Bay NAPS stations for Rule 2 (methodology Section 5): 60807, 60809
THUNDER_BAY_STATION_IDS = {"60807", "60809"}

# Per methodology Section 2.3: include only stations with R >= 0.30
MIN_CORRELATION_R = 0.30

# Research Excel filename patterns (city -> possible filenames to try)
RESEARCH_EXCEL_PATTERNS = {
    "Toronto": ["01. Toronto_Network_All_Regression_Formulas.xlsx"],
    "Montreal": ["01.Montreal_Network_All_Regression_Formulas.xlsx"],
    "Edmonton": ["01.Edmonton_Network_All_Regression_Formulas.xlsx"],
    "Vancouver": ["01.Vancouver_Network_All_Regression_Formulas.xlsx"],
}

CITIES = {
    "Toronto":   {"label": "Toronto",   "lat": 43.7479, "lon": -79.2741},
    "Montreal":  {"label": "Montréal",  "lat": 45.5027, "lon": -73.6639},
    "Edmonton":  {"label": "Edmonton",  "lat": 53.5482, "lon": -113.3681},
    "Vancouver": {"label": "Vancouver", "lat": 49.3686, "lon": -123.2767},
}

DEMO_DATA = {
    "Toronto": {
        "60106": 85.0, "66201": 78.0, "65701": 72.0, "61201": 90.0,
        "60302": 65.0, "65401": 55.0, "60609": 30.0, "360291007": 20.0, "61502": 18.0,
        "60807": 40.0,  # Thunder Bay (Rule 2 demo)
    },
    "Montreal": {
        "54801": 80.0, "52001": 75.0, "50801": 68.0, "500070012": 55.0,
        "500070014": 50.0, "500070007": 45.0, "60106": 70.0, "60302": 40.0,
    },
    "Edmonton": {
        "92801": 90.0, "90302": 75.0, "94401": 65.0, "90304": 70.0,
        "91901": 55.0, "92901": 80.0,
    },
    "Vancouver": {
        "100316": 60.0, "100313": 55.0, "102301": 85.0, "102302": 80.0,
        "100304": 50.0, "100308": 45.0,
    },
}

# Background PM2.5 (µg/m³) for stations not listed in DEMO_DATA (demo / live preview)
DEMO_DEFAULT_PM25 = 14.0

# When using bundled JSON only, pad each city to about this many rows (synthetic upstream sites).
BUNDLED_TARGET_STATION_COUNT = 22

# Cache loaded stations so we don't re-read Excel on every request
_station_cache = {}
_naps_coord_cache = None
_bundled_json_cache = None


def _find_col(headers, *candidates):
    for i, h in enumerate(headers):
        if h is None:
            continue
        hl = str(h).lower().strip()
        for c in candidates:
            if c.lower() in hl:
                return i
    return None


def _get_bundled_json():
    """Parse bundled_stations.json once. Returns dict city_key -> list or None if missing/invalid."""
    global _bundled_json_cache
    if _bundled_json_cache is not None:
        return _bundled_json_cache

    path = Path(__file__).resolve().parent / "bundled_stations.json"
    if not path.is_file():
        _bundled_json_cache = {}
        return _bundled_json_cache

    try:
        with open(path, encoding="utf-8") as f:
            raw = json.load(f)
    except (OSError, json.JSONDecodeError):
        _bundled_json_cache = {}
        return _bundled_json_cache

    if isinstance(raw, list):
        _bundled_json_cache = {}
        return _bundled_json_cache

    if not isinstance(raw, dict):
        _bundled_json_cache = {}
        return _bundled_json_cache

    _bundled_json_cache = raw
    return _bundled_json_cache


def _expand_bundled_stations(city_key, stations):
    """Pad bundled catalog with synthetic corridor stations (coordinates only for WAQI matching)."""
    if not stations or len(stations) >= BUNDLED_TARGET_STATION_COUNT:
        return stations
    prefix = {
        "Toronto": "871",
        "Montreal": "872",
        "Edmonton": "873",
        "Vancouver": "874",
    }.get(city_key, "879")
    expanded = list(stations)
    i = 0
    while len(expanded) < BUNDLED_TARGET_STATION_COUNT:
        base = stations[i % len(stations)]
        nid = f"{prefix}{len(expanded):03d}"
        lat = float(base["lat"]) + 0.14 * (((i % 6) - 3) / 12.0)
        lon = float(base["lon"]) + 0.16 * (((i % 8) - 4) / 12.0)
        expanded.append({
            "id": nid,
            "city_name": base["city_name"],
            "distance": min(620.0, float(base["distance"]) + (i % 6) * 24.0),
            "direction": base["direction"],
            "tier": 2,
            "R": max(0.32, min(0.72, float(base["R"]) - 0.03 * (i % 5))),
            "slope": float(base["slope"]),
            "intercept": float(base["intercept"]),
            "data_type": "synthetic",
            "lat": lat,
            "lon": lon,
        })
        i += 1
    return expanded


def _load_stations_from_bundled(city_key):
    """Load stations from bundled_stations.json for this target city. Returns None if unavailable."""
    data = _get_bundled_json()
    rows = data.get(city_key) if data else None
    if not rows or not isinstance(rows, list):
        return None

    out = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        sid = str(row.get("id", "")).strip()
        if not sid or sid in EXCLUDED_STATION_IDS:
            continue
        try:
            lat = row.get("lat")
            lon = row.get("lon")
            if lat is not None:
                lat = float(lat)
            if lon is not None:
                lon = float(lon)
            out.append({
                "id": sid,
                "city_name": str(row.get("city_name") or ""),
                "distance": float(row.get("distance") or 0),
                "direction": str(row.get("direction") or ""),
                "tier": int(row.get("tier") or 1),
                "R": float(row.get("R") or 0),
                "slope": float(row.get("slope") or 0),
                "intercept": float(row.get("intercept") or 0),
                "data_type": str(row.get("data_type") or ""),
                "lat": lat,
                "lon": lon,
            })
        except (ValueError, TypeError):
            continue

    if not out:
        return None
    out = _expand_bundled_stations(city_key, out)
    return out


def _find_research_excel_path(city_key):
    """Return path to research Excel file, or None if not found."""
    if not RESEARCH_DATA_BASE or not os.path.isdir(RESEARCH_DATA_BASE):
        return None
    city_dir = os.path.join(RESEARCH_DATA_BASE, city_key)
    if not os.path.isdir(city_dir):
        return None
    patterns = RESEARCH_EXCEL_PATTERNS.get(city_key, [f"01.{city_key}_Network_All_Regression_Formulas.xlsx"])
    for pattern in patterns:
        path = os.path.join(city_dir, pattern)
        if os.path.isfile(path):
            return path
    # Fallback: try any xlsx file starting with 01
    for f in os.listdir(city_dir):
        if f.startswith("01") and f.endswith(".xlsx"):
            return os.path.join(city_dir, f)
    return None


def _load_stations_from_research(city_key):
    """
    Load stations from research Excel format.
    Sheet: first sheet (e.g. Toronto_ALL_Regression_ON_QC)
    Headers: row 4 (0-indexed)
    Columns: Station ID, City, Country, Distance (km), Direction, Tier, Data Type, Regression Formula, R, R², Slope, Intercept, ...
    """
    path = _find_research_excel_path(city_key)
    if not path:
        return None

    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception:
        return None

    if len(rows) < 6:
        return None

    headers = [str(h).strip() if h else "" for h in rows[4]]
    col_id = _find_col(headers, "station id")
    col_city = _find_col(headers, "city")
    col_dist = _find_col(headers, "distance")
    col_dir = _find_col(headers, "direction")
    col_tier = _find_col(headers, "tier")
    col_slope = _find_col(headers, "slope")
    col_int = _find_col(headers, "intercept")
    col_dtype = _find_col(headers, "data type")

    col_r = None
    for i, h in enumerate(headers):
        if h.strip() in ("R", "R²"):
            col_r = i
            break

    if col_id is None:
        return None

    stations = []
    for row in rows[5:]:
        if row[col_id] is None:
            continue
        sid = str(row[col_id]).strip()
        if not sid:
            continue
        # Skip header/label rows (e.g. "Rule 3 Québec Stations", "Legend:")
        sid_clean = sid.replace(" ", "").replace(".", "").replace("-", "")
        if not sid_clean.isdigit() or len(sid) > 15:
            continue
        if sid in EXCLUDED_STATION_IDS:
            continue
        # Per methodology Section 2.3: R >= 0.30
        r_val = row[col_r] if col_r is not None and row[col_r] else 0
        try:
            if float(r_val) < MIN_CORRELATION_R:
                continue
        except (ValueError, TypeError):
            continue
        city_name = str(row[col_city] or "") if col_city is not None else ""
        try:
            dist_val = row[col_dist] if col_dist is not None else 0
            tier_val = row[col_tier] if col_tier is not None else 1
            r_val = row[col_r] if col_r is not None and row[col_r] else 0
            slope_val = row[col_slope] if col_slope is not None else 0
            int_val = row[col_int] if col_int is not None else 0
            stations.append({
                "id": sid,
                "city_name": city_name,
                "distance": float(dist_val) if dist_val else 0,
                "direction": str(row[col_dir] or "") if col_dir is not None else "",
                "tier": int(str(tier_val).replace("Tier", "").strip()) if tier_val else 1,
                "R": float(r_val) if r_val else 0,
                "slope": float(slope_val) if slope_val else 0,
                "intercept": float(int_val) if int_val else 0,
                "data_type": str(row[col_dtype] or "") if col_dtype is not None else "",
            })
        except (ValueError, TypeError):
            continue

    return stations


def _load_stations_from_legacy(city_key):
    """Load stations from legacy format (Included Stations sheet)."""
    fn = os.path.join(DATA_DIR, f"{city_key}_PM25_EWS_Regression.xlsx")
    if not os.path.exists(fn):
        return None

    try:
        import openpyxl
        wb = openpyxl.load_workbook(fn, read_only=True, data_only=True)
        ws = wb["Included Stations"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception:
        return None

    if len(rows) < 3:
        return None

    headers = [str(h).strip() if h else "" for h in rows[1]]
    col_id = _find_col(headers, "station id")
    col_city = _find_col(headers, "city")
    col_dist = _find_col(headers, "distance")
    col_dir = _find_col(headers, "direction")
    col_tier = _find_col(headers, "tier")
    col_slope = _find_col(headers, "slope")
    col_int = _find_col(headers, "intercept")
    col_dtype = _find_col(headers, "data type")

    col_r = None
    for i, h in enumerate(headers):
        if h.strip() == "R":
            col_r = i
            break

    if col_id is None:
        return None

    stations = []
    for row in rows[2:]:
        if row[col_id] is None:
            continue
        sid = str(row[col_id]).strip()
        if not sid:
            continue
        if sid in EXCLUDED_STATION_IDS:
            continue
        city_name = str(row[col_city] or "")
        try:
            stations.append({
                "id": sid,
                "city_name": city_name,
                "distance": float(row[col_dist]) if row[col_dist] else 0,
                "direction": str(row[col_dir] or ""),
                "tier": int(str(row[col_tier]).replace("Tier", "").strip()) if row[col_tier] else 1,
                "R": float(row[col_r]) if col_r is not None and row[col_r] else 0,
                "slope": float(row[col_slope]) if row[col_slope] else 0,
                "intercept": float(row[col_int]) if row[col_int] else 0,
                "data_type": str(row[col_dtype] or "") if col_dtype is not None else "",
            })
        except (ValueError, TypeError):
            continue

    return stations


def _load_naps_coords():
    """Load NAPS station coordinates. Returns {station_id: (lat, lon)}."""
    global _naps_coord_cache
    if _naps_coord_cache is not None:
        return _naps_coord_cache

    if not NAPS_STATIONS_PATH or not os.path.isfile(NAPS_STATIONS_PATH):
        _naps_coord_cache = {}
        return _naps_coord_cache

    try:
        import openpyxl
        wb = openpyxl.load_workbook(NAPS_STATIONS_PATH, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception:
        _naps_coord_cache = {}
        return _naps_coord_cache

    if len(rows) < 2:
        _naps_coord_cache = {}
        return _naps_coord_cache

    headers = [str(h).strip() if h else "" for h in rows[0]]
    col_id = _find_col(headers, "naps id")
    col_lat = _find_col(headers, "latitude")
    col_lon = _find_col(headers, "longitude")

    if col_id is None or col_lat is None or col_lon is None:
        _naps_coord_cache = {}
        return _naps_coord_cache

    coords = {}
    for row in rows[1:]:
        if row[col_id] is None:
            continue
        sid = str(row[col_id]).strip()
        try:
            lat = float(row[col_lat])
            lon = float(row[col_lon])
            coords[sid] = (lat, lon)
        except (ValueError, TypeError):
            continue

    _naps_coord_cache = coords
    return coords


def _load_coords_legacy(city_key):
    """Load lat/lon from 'All Stations Data' sheet. Returns {station_id: (lat, lon)}."""
    fn = os.path.join(DATA_DIR, f"{city_key}_PM25_EWS_Regression.xlsx")
    if not os.path.exists(fn):
        return {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(fn, read_only=True, data_only=True)
        ws = wb["All Stations Data"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception:
        return {}

    if len(rows) < 3:
        return {}

    headers = [str(h).strip() if h else "" for h in rows[1]]
    col_id = _find_col(headers, "station id")
    col_lat = _find_col(headers, "lat")
    col_lon = _find_col(headers, "lon")

    if col_id is None or col_lat is None or col_lon is None:
        return {}

    coords = {}
    for row in rows[2:]:
        if row[col_id] is None:
            continue
        sid = str(row[col_id]).strip()
        try:
            lat = float(row[col_lat])
            lon = float(row[col_lon])
            coords[sid] = (lat, lon)
        except (ValueError, TypeError):
            continue
    return coords


def _load_thunder_bay_rule2_stations():
    """Thunder Bay stations for Rule 2 (methodology Section 5). Weak correlation with Toronto;
    used only for trigger (>35 µg/m³), not for prediction. Coords from NAPS."""
    naps = _load_naps_coords()
    stations = []
    for sid in THUNDER_BAY_STATION_IDS:
        c = naps.get(sid)
        if not c:
            continue
        # ~1200 km NW of Toronto; dummy regression (trigger-only, methodology says weak R)
        stations.append({
            "id": sid,
            "city_name": "Thunder Bay",
            "distance": 1200.0,
            "direction": "NW",
            "tier": 2,
            "R": 0.0,
            "slope": 0.3,
            "intercept": 5.0,
            "data_type": "Rule2",
            "lat": c[0],
            "lon": c[1],
        })
    return stations


def _get_coord_map(city_key, from_research):
    """Get coordinate map for stations. Tries legacy sheet first, then NAPS lookup."""
    coord_map = _load_coords_legacy(city_key)
    if coord_map:
        return coord_map
    naps = _load_naps_coords()
    return naps


def load_stations(city_key):
    if city_key in _station_cache:
        return _station_cache[city_key]

    stations = _load_stations_from_research(city_key)
    from_research = stations is not None
    if stations is None:
        stations = _load_stations_from_legacy(city_key)
        if stations is not None:
            from_research = False

    if stations is None:
        stations = []

    if len(stations) == 0:
        bundled = _load_stations_from_bundled(city_key)
        if bundled:
            stations = bundled
            from_research = False

    if len(stations) == 0:
        _station_cache[city_key] = []
        return []

    if city_key == "Toronto":
        existing_ids = {st["id"] for st in stations}
        tb_stations = _load_thunder_bay_rule2_stations()
        stations = stations + [st for st in tb_stations if st["id"] not in existing_ids]

    coord_map = _get_coord_map(city_key, from_research)
    for st in stations:
        c = coord_map.get(st["id"])
        if c:
            st["lat"], st["lon"] = c[0], c[1]
        elif st.get("lat") is None or st.get("lon") is None:
            st["lat"] = None
            st["lon"] = None

    stations.sort(key=lambda s: (s["tier"], -s["distance"]))
    _station_cache[city_key] = stations
    return stations


def load_all_stations():
    """Load stations from all cities, tagging each with its target city."""
    if "_all" in _station_cache:
        return _station_cache["_all"]
    all_stations = []
    for city_key in CITIES:
        for st in load_stations(city_key):
            st_copy = dict(st)
            st_copy["target_city"] = city_key
            all_stations.append(st_copy)
    all_stations.sort(key=lambda s: (s["target_city"], s["tier"], -s["distance"]))
    _station_cache["_all"] = all_stations
    return all_stations


def get_all_demo_data():
    """Merge demo data from all cities into one dict."""
    merged = {}
    for city_data in DEMO_DATA.values():
        merged.update(city_data)
    return merged


def build_demo_previous_readings(stations):
    """Flat id -> pm for Rule 2 continuity in demo/preview (last city wins on duplicate ids)."""
    prev = {}
    for st in stations:
        tc = st.get("target_city", "")
        sid = st["id"]
        city_demo = DEMO_DATA.get(tc, {})
        prev[sid] = float(city_demo[sid]) if sid in city_demo else float(DEMO_DEFAULT_PM25)
    return prev


def _load_coords(city_key):
    """Legacy: Load lat/lon. Returns {station_id: (lat, lon)}."""
    return _get_coord_map(city_key, from_research=False)
